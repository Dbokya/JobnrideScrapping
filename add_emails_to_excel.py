import openpyxl
from openpyxl import load_workbook

# Read emails from useraddfile.txt
with open('linkedin_email_campaign/useraddfile.txt', 'r', encoding='utf-8') as f:
    new_emails = [line.strip() for line in f if line.strip()]

print(f"Found {len(new_emails)} emails to add")

# Load the existing Excel file
wb = load_workbook('linkedin_email_campaign/users.xlsx')
ws = wb.active

print(f"Current sheet: {ws.title}")
print(f"Current rows: {ws.max_row}")

# Read existing emails (assuming emails are in column A)
existing_emails = []
for row in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=1).value
    if cell_value:
        existing_emails.append(cell_value)

print(f"Found {len(existing_emails)} existing emails")

# Insert new rows at the top
ws.insert_rows(1, len(new_emails))

# Add new emails to the top
for idx, email in enumerate(new_emails, start=1):
    ws.cell(row=idx, column=1, value=email)

# Save the workbook
wb.save('linkedin_email_campaign/users.xlsx')

print(f"\nSuccessfully added {len(new_emails)} emails to the top of users.xlsx")
print(f"Total rows now: {ws.max_row}")
